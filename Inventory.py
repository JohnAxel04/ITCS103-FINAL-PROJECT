import tkinter as tk
import openpyxl as op
from tkinter import ttk,messagebox
import datetime
from datetime import datetime
def valid():
    entry = productEntry.get()
    qty = productquatity.get()
    price = productPrice.get()
    exp = productExpiry.get()

    if not entry or not qty or not price or not exp:
        messagebox.showerror("Invalid Input","Input must not be empty")
        return False
    if not qty.isdigit() or not price.isdigit():
        messagebox.showerror("Invalid Input","Quantity and Price must be numbers")
        return False
    
    try:
        expiry_date = datetime.strptime(exp, "%Y-%m-%d",).date()
        today = datetime.today().date()

        if expiry_date < today:
            messagebox.showerror("Expired","Item already expired")
            return False
        
    except:
        messagebox.showerror("Error","User format YYYY-MM-DD")
        return False

    return True

def save():
    if not valid():
        return
    entry = productEntry.get()
    qty = int(productquatity.get())
    price = int(productPrice.get())
    exp = productExpiry.get()
    opt = typeOption.get()
    wrk = op.load_workbook("Inventory.xlsx")
    sheet = wrk.active
    totalprice = qty * price

    oldid = sheet.cell(row=sheet.max_row, column=1).value
    id = oldid + 1

    sheet.append([id,entry,qty,price,exp,opt,totalprice])
    wrk.save("Inventory.xlsx")
    messagebox.showinfo("Successfull","Item added successfully")
    productEntry.delete(0,tk.END)
    productquatity.delete(0,tk.END)
    productPrice.delete(0,tk.END)
    productExpiry.delete(0,tk.END)
    typeOption.current(0)
    refresh()
def refresh():
    work = op.load_workbook("Inventory.xlsx")
    sheet = work.active
    for i in table.get_children():
        table.delete(i)
    
    for row in sheet.iter_rows(min_row=2,values_only=True):
        packaging_type = row[5]
        table.insert("",tk.END,values=row,tags=(packaging_type,))
def focus(event):
    selected = table.focus()
    values = table.item(selected,"values")
    
    if values:
        productEntry.delete(0,tk.END)
        productquatity.delete(0,tk.END)
        productPrice.delete(0,tk.END)
        productExpiry.delete(0,tk.END)
        typeOption.current(0)

        typeopt = values[5]
        productEntry.insert(0,values[1])
        productquatity.insert(0,values[2])
        productPrice.insert(0,values[3])
        productExpiry.insert(0,values[4])
        typeOption.set(typeopt)

def update():
    select = table.focus()
    values = table.item(select,"values")

    if not select:
        messagebox.showerror("error","Select to update")
        return
    if not valid():
        return
    
    id = values[0]
    entry = productEntry.get()
    qty = int(productquatity.get())
    price = int(productPrice.get())
    exp = productExpiry.get()
    opt = typeOption.get()
    total = qty * price

    work = op.load_workbook("Inventory.xlsx")
    sheet = work.active
    for i in sheet.iter_rows(min_row=2):
        if str(i[0].value) == str(id):
            i[1].value = entry
            i[2].value = qty
            i[3].value = price
            i[4].value = exp
            i[5].value = opt
            i[6].value = total
    work.save("Inventory.xlsx")
    messagebox.showinfo("Success","Info successfully updated")
    refresh()
def delete():
    select = table.focus()
    values = table.item(select,"values")

    if not select:
        messagebox.showerror("error","Select to update")
        return
    
    confirm = messagebox.askyesnocancel("Delete","Want to delete?")
    if not confirm:
        return
    id = values[0]
    work = op.load_workbook("Inventory.xlsx")
    sheet = work.active
    for i,row in enumerate(sheet.iter_rows(min_row=2),start=2):
        if str(row[0].value) == str(id):
            sheet.delete_rows(i)

    productEntry.delete(0,tk.END)
    productquatity.delete(0,tk.END)
    productPrice.delete(0,tk.END)
    productExpiry.delete(0,tk.END)
    typeOption.current(0)
    work.save("Inventory.xlsx")
    messagebox.showinfo("Delete","Info successfully deleted")
    refresh()

def search():
    searched = searchEntry.get().lower()

    work = op.load_workbook("Inventory.xlsx")
    sheet = work.active

    for i in table.get_children():
        table.delete(i)
    
    for row in sheet.iter_rows(min_row=2,values_only=True):
        name = str(row[1]).lower()
        types = str(row[5]).lower()

        if searched in name or searched in types:
            table.insert("",tk.END,values=row)


window = tk.Tk()
window.title("Inventory System")
mainlabel = tk.Label(window,text="Inventory Management System",font=("Poppins",15,"bold"))
mainlabel.grid(columnspan=5,pady=10)
frame = tk.Frame(window,background="#888")
frame.grid(row=1,columnspan=6,padx=5,pady=8)

productEntry = tk.Entry(frame)
productEntry.grid(row=0,column=0,padx=5,pady=(8,2))
productquatity = tk.Entry(frame)
productquatity.grid(row=0,column=1,padx=5,pady=(8,2))
productPrice = tk.Entry(frame)
productPrice.grid(row=0,column=2,padx=5,pady=(8,2))
productExpiry = tk.Entry(frame)
productExpiry.grid(row=0,column=3,padx=5,pady=(8,2))
options = ['Box','Pack','Sachet',"Galon"]
typeOption = ttk.Combobox(frame,values=options,state="readonly")
typeOption.current(0)
typeOption.grid(column=4,row=0,padx=5,pady=(8,2))
searchEntry = tk.Entry(frame)
searchEntry.grid(column=0,row=2,columnspan=2,sticky="ew",padx=(5,0))
searchLabel = tk.Label(frame,text="Search",background="#888",foreground="white")
searchLabel.grid(column=0,row=3,columnspan=2)
searchbtn = tk.Button(frame,text="Search",height=1,width=8,command=search,activebackground="#e67e22",background="#f39c12",fg="white",font=("poppins",9,"bold"))
searchbtn.grid(column=2,row=2,sticky="w",padx=5)

NameLabel = tk.Label(frame,text="Product Name",background="#888",foreground="white")
NameLabel.grid(row=1,column=0)
QuantityLabel = tk.Label(frame,text="Quantity",background="#888",foreground="white") 
QuantityLabel.grid(row=1,column=1) 
PriceLabel = tk.Label(frame,text="Price",background="#888",foreground="white")
PriceLabel.grid(row=1,column=2)
ExpiryLabel = tk.Label(frame,text="Expiry Date",background="#888",foreground="white")
ExpiryLabel.grid(row=1,column=3)
TypeLabel = tk.Label(frame,text="Type of Packaging",background="#888",foreground="white")
TypeLabel.grid(row=1,column=4)
def savebt(event):
    SaveBtn['bg'] = "#2980b9"
def outsave(event):
    SaveBtn['bg'] = "#3498db"
btnframe = tk.Frame(window)
btnframe.grid(row=2,columnspan=4)
SaveBtn = tk.Button(btnframe,text="Submit",command=save,activebackground="#2980b9",background="#3498db",fg="white",width=9,height=2,font=("poppins",9,"bold"))
SaveBtn.grid(column=1,row=0,padx=10,pady=(1,10))
SaveBtn.bind("<Enter>",savebt)
SaveBtn.bind("<Leave>",outsave)

def updbt(event):
    UpdateBtn['bg'] = "#27ae60"
def outup(event):
    UpdateBtn['bg'] = "#2ecc71"
UpdateBtn = tk.Button(btnframe,text="Update",command=update,activebackground="#27ae60",background="#2ecc71",fg="white",width=8,height=1,font=("poppins",9,"bold"))
UpdateBtn.grid(column=0,row=0,padx=10,pady=(1,10))

UpdateBtn.bind("<Enter>",updbt)
UpdateBtn.bind("<Leave>",outup)

def updel(event):
    DeleteBtn['bg'] = "#c0392b"
def outdel(event):
    DeleteBtn['bg'] = "#e74c3c"
DeleteBtn = tk.Button(btnframe,text="Delete",command=delete,activebackground="#c0392b",background="#e74c3c",fg="white",width=8,height=1,font=("poppins",9,"bold"))
DeleteBtn.grid(column=2,row=0,padx=10,pady=(1,10))

DeleteBtn.bind("<Enter>",updel)
DeleteBtn.bind("<Leave>",outdel)

style = ttk.Style()

style.theme_use("clam")

style.configure("Treeview",backgound = "white",foreground = "black",rowheight = 25,fieldbackground = "white",font = ("Poppins",10))

style.configure("Treeview.Heading",background = "#8492a0",foreground = "white",font=("Poppins", 10, "bold"))

style.map( "Treeview",background=[("selected", "#3498db")],foreground=[("selected", "white")])


table = ttk.Treeview(window,columns=("ID","Product Name","Quantity","Price","Expiry","Type of packaging","Total Price"),show="headings")
for row in ("ID","Product Name","Quantity","Price","Expiry","Type of packaging","Total Price"):
    table.heading(row, text=row)

table.column("ID", width=50, anchor="center")
table.column("Product Name", width=180)
table.column("Quantity", width=100, anchor="center")
table.column("Price", width=100, anchor="center")
table.column("Expiry", width=120, anchor="center")
table.column("Type of packaging", width=150, anchor="center")
table.column("Total Price", width=120, anchor="center")

table.tag_configure("Box", background="#d6eaf8")
table.tag_configure("Pack", background="#d5f5e3")
table.tag_configure("Sachet", background="#fcf3cf")
table.tag_configure("Galon", background="#fadbd8")

table.grid(columnspan=4,row=3,pady=(0,20),padx=20)
table.bind("<<TreeviewSelect>>",focus)
refresh()

footer = tk.Frame(window,background="darkgrey")
footer.grid(column=0,row=4,columnspan=4,sticky="we")
footerLabel = tk.Label(footer,text="@johnaxel04 5-2026",background="darkgrey")
footerLabel.pack()
window.mainloop()